# -*- coding: utf-8 -*-
"""
Created on Wed Sep 22 14:50:09 2021

@author: Administrator
"""
def wimi_zzgs(): 
    "WIMI最终格式"
    
    import akshare as ak
    import pandas as pd
    import time 
    from matplotlib import pyplot as plt 
    import numpy as np
    import datetime
    from dateutil.relativedelta import relativedelta
    from openpyxl import load_workbook
    import xlsxwriter
    import openpyxl
    from openpyxl.styles import PatternFill,Font,Alignment
    
    
    def getLastWeekDay(day=datetime.datetime.now()):
        now=day
        if now.isoweekday()==1:
          dayStep=3
        else:
          dayStep=1
        lastWorkDay = now - datetime.timedelta(days=dayStep)
        return lastWorkDay
    date_work=getLastWeekDay().strftime('%Y%m%d')
    
    
    
    df_info = pd.DataFrame(columns=('idd','ftnn_bs'))
    df_info.loc[0,'idd'] = 'WIMI'
    df_info.loc[1,'idd'] = 'MLGO'
    df_info.loc[2,'idd'] = 'HOLO'
    # df_info.loc[3,'idd'] = 'VENAR'
    
    
    def zzgs(idd):   #数据整合到一个表格中
        
        ans = pd.read_excel('{}笔数&股数{}.xlsx'.format(idd,date_work),engine='openpyxl')[['时间','每分钟笔数','每5分钟笔数','每分钟股数','每5分钟股数','每分钟笔数(通达信)','每5分钟笔数(通达信)']]
        
        
        "以下为最终放到同一个sheet中的数据"
        
        ans_5m = ans.drop(ans[np.isnan(ans['每5分钟股数'])].index).reset_index(drop=True)[['时间','每5分钟股数','每5分钟笔数','每5分钟笔数(通达信)']].rename(columns={'时间':'盘中','每5分钟股数':'股数','每5分钟笔数':'笔数（同花顺版本）','每5分钟笔数(通达信)':'笔数(通达信)'}).set_index(["盘中"])                                        
        # ans_5m['笔数（同花顺版本）'] = round(ans_5m['笔数（同花顺版本）'],0)
        
        ans_gs = pd.read_excel('{}{}--股数.xlsx'.format(idd,date_work),engine='openpyxl',dtype={'时间':'str'}).iloc[:,[0,1,2,4]].set_index(["时间"])
        
        ans_1m_gs = ans.drop(ans[np.isnan(ans['每分钟股数'])].index).reset_index(drop=True)[['时间','每分钟股数']].rename(columns={'每分钟股数':'股数'}).set_index(["时间"]) 
        
        "存放到同一个sheet中"    
        def to_onesheet(file_name, sheet_name, df_list, direction, spaces):
            row = 1
            col = 0
            writer = pd.ExcelWriter(file_name)
            
            # writer = pd.ExcelWriter('美股笔数股数{}.xlsx'.format(date_work))
            
            
            
            for dataframe in df_list:
                dataframe.to_excel(writer, sheet_name,
                                   startrow=row, startcol=col)
                if direction == 'h':
                    col = col + len(dataframe.columns) + spaces + 1
                elif direction == 'v':
                    row = row + len(dataframe.index) + spaces + 1
                else:
                    raise ValueError(
                        f"Direction must be 'h' or 'v', you entered is '{direction}'")   
                
              
            return writer.save()
        
        
        df_list = [ans_gs,ans_1m_gs,ans_5m]
        
        to_onesheet('{}{}.xlsx'.format(idd,date_work), '{}'.format(idd), df_list, 'h', 1)
        
    
        
        "在具体位置上添加字符"
        workbook=openpyxl.load_workbook("{}{}.xlsx".format(idd,date_work))
        worksheet=workbook.worksheets[0]     # 从0开始
        
        worksheet["A1"]="原始数据（股数正确版本）"
        worksheet["F1"]="每隔一分钟（股数正确版本）"
        worksheet["I1"]="每5分钟（股数正确版本）"
        worksheet["I83"] = "求和"
        
        "字体加粗"
        worksheet["A1"].font = Font(bold=True)
        worksheet["F1"].font = Font(bold=True)
        worksheet["I1"].font = Font(bold=True)
        worksheet["I83"].font = Font(bold=True)
        worksheet["J83"].font = Font(bold=True)
        worksheet["K83"].font = Font(bold=True)
        
        "填充颜色"
        
        fille=PatternFill('solid', fgColor='FFBB02')
        for i in range(1,12):
            worksheet.cell(row=1,column=i).fill = fille
            
        fille=PatternFill('solid', fgColor='FFFF00')
        for i in range(3,82):
            for j in range(10,13):
                worksheet.cell(row=i,column=j).fill = fille
        
        
        "股数笔数反常识标红----修改日期20221012"
        fille=PatternFill('solid', fgColor='FF0000')
        for i in range(3,82):
            if  (round(worksheet['K{}'.format(i)].value)*100>worksheet['J{}'.format(i)].value) or ((round(worksheet['K{}'.format(i)].value)==0)&(worksheet['J{}'.format(i)].value!=0)) :
                worksheet.cell(row=i,column=11).fill = fille
        
        
        "插入求和公式"  
        worksheet["J83"] = "=SUM(J2:J81)"
        worksheet["K83"] = "=SUM(K2:K81)"
        fille=PatternFill('solid', fgColor='FFA500')
        worksheet.cell(row=83,column=10).fill = fille
        worksheet.cell(row=83,column=11).fill = fille
        
        
        
        "设置列宽"
        # wb=openpyxl.Workbook()
        # sheet=wb.active
        worksheet.column_dimensions['A'].width=15
        worksheet.column_dimensions['F'].width=15
        worksheet.column_dimensions['I'].width=15
        
        
        "笔数（同花顺版本）   K2单元格左对齐"
        worksheet['K2'].alignment = Alignment(horizontal='left',vertical='center',wrap_text=False)
        "每5分钟股数只显示整数"
        for i in range(3,84):
            worksheet["K{}".format(i)].number_format = '###'
                
        
        workbook.save(filename="{}{}.xlsx".format(idd,date_work))
        
        print(idd)
    
    
    
    for index,row in df_info.iterrows():
        zzgs(row['idd'])






if __name__ == '__main__':
    wimi_zzgs()
    pass    


























