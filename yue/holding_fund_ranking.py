# -*- coding: utf-8 -*-
"""
Created on Thu Feb 23 14:51:57 2023

@author: Administrator
"""


def holding_fund_ranking_1():   
    import akshare as ak
    import pandas as pd
    import time 
    from matplotlib import pyplot as plt 
    import numpy as np
    import datetime
    from dateutil.relativedelta import relativedelta
    from openpyxl import load_workbook
    import xlsxwriter
    import openpyxl
    from openpyxl.styles import PatternFill,Font,Alignment
    from functools import reduce
    import statsmodels.api as sm
    # from itertools import *
    import math
    time_start = time.time()
    
    def getLastWeekDay(day=datetime.datetime.now()):
        now=day
        if now.isoweekday()==1:
          dayStep=3
        else:
          dayStep=1
        lastWorkDay = now - datetime.timedelta(days=dayStep)
        return lastWorkDay
    date_work=getLastWeekDay().strftime('%Y%m%d')
    now = datetime.datetime.now().strftime('%Y%m%d')
    def bu_zero(fund_id,id_id):
        "基金代码前面补0"
        "fund_id：df名称  id_id:需要修改的列"
        fund_id['{}'.format(id_id)] = fund_id['{}'.format(id_id)].map(lambda x:str(x).zfill(6))
        return fund_id
    
    
    
    
    s_fund = ak.fund_open_fund_rank_em(symbol="股票型")
    m_fund = ak.fund_open_fund_rank_em(symbol="混合型")
    i_fund = ak.fund_open_fund_rank_em(symbol="指数型")
    q_fund = ak.fund_open_fund_rank_em(symbol="QDII")
    
    
    
    fund_all = pd.concat([s_fund,m_fund])[['基金代码','基金简称','日期','近1周','近1月','近3月','近6月','近1年','近2年']].copy()  #股票+混合
    
    fund_df = pd.read_excel('基金持仓名单.xlsx',engine='openpyxl',dtype={'code':'str'}) 
    fund_df = bu_zero(fund_df,'基金代码') 
    
    
    fund_df_index = pd.merge(fund_df,i_fund[['基金代码','基金简称','日期','近1周','近1月','近3月','近6月','近1年','近2年']].copy(),on='基金代码',how='inner')
    fund_all = pd.concat([fund_all,fund_df_index])  #股票+混合+ETF
    
    
    fund_df_qdii = pd.merge(fund_df,q_fund[['基金代码','基金简称','日期','近1周','近1月','近3月','近6月','近1年','近2年']].copy(),on='基金代码',how='inner')
    fund_all = pd.concat([fund_all,fund_df_qdii])  #股票+混合+ETF+QDII
    
    
    
    fund_all = fund_all.drop_duplicates().reset_index(drop=True)
    fund_df = pd.merge(fund_df,fund_all[['基金代码','基金简称']].copy(),on='基金代码',how='left')
    
    
    fund_all = fund_all.replace('',np.nan)
    fund_all.iloc[:,3:] = fund_all.iloc[:,3:].astype('float')
    
    
    
    cl_list = ['近1周','近1月','近3月','近6月','近1年','近2年']
    for i in cl_list[:]:
        data_b = fund_all.sort_values(i,ascending=0).reset_index(drop=True).reset_index()
        data_b['{}'.format(i)] =  data_b['index']+1
        fund_df = pd.merge(fund_df,data_b[['基金代码','{}'.format(i)]].copy(),how='left',on='基金代码')
    fund_df.columns = ['基金代码','基金简称','近1周排名','近1月排名','近3月排名','近6月排名','近1年排名','近2年排名']
    
    fund_df = pd.merge(fund_df,fund_all[['基金代码','近1周','近1月','近3月','近6月','近1年','近2年']].copy(),how='left',on='基金代码')
    
    fund_df.to_excel('持仓基金排名{}.xlsx'.format(date_work))
    
    
    
    "小于500  高亮浅红色"
    workbook=openpyxl.load_workbook("持仓基金排名{}.xlsx".format(date_work))
    worksheet=workbook.worksheets[0]     # 从0开始
    
    fille=PatternFill('solid', fgColor='FFC0CB')
    n=4
    for j in ['D','E','F','G','H','I']:
        for i in range(2,len(fund_df)+2):
            if worksheet['{}{}'.format(j,i)].value<=500:
                worksheet.cell(row=i,column=n).fill = fille
        n+=1
    
    workbook.save(filename="持仓基金排名{}.xlsx".format(date_work))
    
    
if __name__ == '__main__':
    holding_fund_ranking_1()
    pass











