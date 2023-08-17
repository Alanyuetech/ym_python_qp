# -*- coding: utf-8 -*-
"""
Created on Mon Aug 30 14:01:39 2021

@author: Administrator
"""

"A股实时"
import akshare as ak
import pandas as pd
import time 
from matplotlib import pyplot as plt 
import numpy as np
import datetime
from dateutil.relativedelta import relativedelta
from openpyxl import load_workbook
import xlsxwriter
import openpyxl
from openpyxl.styles import PatternFill,Font,Alignment

now_day = datetime.datetime.now().strftime('%Y%m%d')
now_time = datetime.datetime.now().strftime('%H:%M')

cffex_text = ak.match_main_contract(exchange="cffex")  #金融期货代码获取
df = pd.DataFrame(columns=(['时间','上证成交额前20上涨数量','上证成交额前20平均涨跌幅','深证成交额前20上涨数量','深证成交额前20平均涨跌幅','沪深300主力','上证50主力','中证500主力']))
df.to_excel("成交额前20指标明细{}.xlsx".format(now_day),index=False)

"下午/重新跑 只跑下面的代码"

def sleep_time(hour,minu,sec):
    return hour*3600+minu*60+sec
second = sleep_time(0,0,5)




def a_stock():
    "A股所有股票数据"
    a_stock = ak.stock_zh_a_spot_em()
    
    
    "获取A股实时行情数据——成交额前20"
    a_stock_hs = a_stock[['代码','名称','最新价','涨跌幅','成交量','成交额','换手率']].sort_values(by='成交额',ascending=0).head(20).reset_index(drop=True)
    a_stock_hs.loc[a_stock_hs['涨跌幅'] >= 10,'涨跌幅'] = 10
    a_stock_hs_zdf_go = len(a_stock_hs[a_stock_hs['涨跌幅']>0])  
    print("A股")
    # print(a_stock_hs)
    print("")
    print("A股成交额前20上涨股票数量：{}  平均涨跌幅：{:.2}".format(a_stock_hs_zdf_go,a_stock_hs['涨跌幅'].mean()))
    print("") 
    
    "获取上证实时行情数据——成交额前20"
    a_stock_sh = a_stock[a_stock['代码'].str.slice(0,1)=='6'][['代码','名称','最新价','涨跌幅','成交量','成交额','换手率']].sort_values(by='成交额',ascending=0).head(20).reset_index(drop=True)
    a_stock_sh.loc[a_stock_sh['涨跌幅'] >= 10,'涨跌幅'] = 10
    a_stock_sh_zdf_go = len(a_stock_sh[a_stock_sh['涨跌幅']>0])
    "获取股指期货主力"
    data = ak.futures_zh_spot(subscribe_list=cffex_text, market="FF", adjust=False)
    
    print("上证")
    # print(a_stock_sh)
    time_now = datetime.datetime.now().strftime('%Y.%m.%d-%H:%M:%S')
    print(time_now)
    print("")
    print("上证成交额前20上涨股票数量：{}  平均涨跌幅：{:.2}".format(a_stock_sh_zdf_go,a_stock_sh['涨跌幅'].mean()))
    print("")

    "获取深证实时行情数据——成交额前20"
    a_stock_sz = a_stock[a_stock['代码'].str.slice(0,1)=='0'][['代码','名称','最新价','涨跌幅','成交量','成交额','换手率']].sort_values(by='成交额',ascending=0).head(20).reset_index(drop=True)
    a_stock_sz.loc[a_stock_sz['涨跌幅'] >= 10,'涨跌幅'] = 10
    a_stock_sz_zdf_go = len(a_stock_sz[a_stock_sz['涨跌幅']>0])
    print("深证")   
    # print(a_stock_sz)
    print("")
    print("深证成交额前20上涨股票数量：{}  平均涨跌幅：{:.2}".format(a_stock_sz_zdf_go,a_stock_sz['涨跌幅'].mean()))
    print("")

    print("~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~")
    
    df_b = pd.DataFrame(columns=(['时间','上证成交额前20上涨数量','上证成交额前20平均涨跌幅','深证成交额前20上涨数量','深证成交额前20平均涨跌幅','沪深300主力','上证50主力','中证500主力']))
    df_b.loc[0,'时间'] = time_now
    df_b.loc[0,'上证成交额前20上涨数量'] = "{}".format(a_stock_sh_zdf_go)
    df_b.loc[0,'上证成交额前20平均涨跌幅'] = a_stock_sh['涨跌幅'].mean()
    df_b.loc[0,'深证成交额前20上涨数量'] = "{}".format(a_stock_sz_zdf_go)
    df_b.loc[0,'深证成交额前20平均涨跌幅'] = a_stock_sz['涨跌幅'].mean()
        
    df_b['沪深300主力'] = data.loc[data['symbol'].str.contains('沪深300')].reset_index(drop=True).loc[0,'current_price']
    df_b['上证50主力'] = data.loc[data['symbol'].str.contains('上证50')].reset_index(drop=True).loc[0,'current_price']
    df_b['中证500主力'] = data.loc[data['symbol'].str.contains('中证500')].reset_index(drop=True).loc[0,'current_price']
        
    return df_b



def write_excel(df_b):
    workbook=openpyxl.load_workbook("成交额前20指标明细{}.xlsx".format(now_day))
    worksheet=workbook.worksheets[0]
    worksheet.append([df_b.loc[0,'时间'],df_b.loc[0,'上证成交额前20上涨数量'],df_b.loc[0,'上证成交额前20平均涨跌幅'],df_b.loc[0,'深证成交额前20上涨数量'],df_b.loc[0,'深证成交额前20平均涨跌幅'],df_b.loc[0,'沪深300主力'],df_b.loc[0,'上证50主力'],df_b.loc[0,'中证500主力']])
    workbook.save("成交额前20指标明细{}.xlsx".format(now_day))




"上午"

now_time = datetime.datetime.now().strftime('%H:%M:%S')

while now_time <= '09:29:40':     #     09:29:40      12:59:40
    time.sleep(second)
    now_time = datetime.datetime.now().strftime('%H:%M:%S')
    print(now_time)


while now_time <= '11:30:30':
    time.sleep(second)
    df_b = a_stock()    
    write_excel(df_b)
    df = df.append(df_b)
    now_time = datetime.datetime.now().strftime('%H:%M:%S')



#########################################
"下午"


now_time = datetime.datetime.now().strftime('%H:%M:%S')

while now_time <= '12:59:40':     #     09:29:40      12:59:40
    time.sleep(second)
    now_time = datetime.datetime.now().strftime('%H:%M:%S')
    print(now_time)


while now_time <= '15:00:30':
    time.sleep(second)
    df_b = a_stock()    
    write_excel(df_b)
    df = df.append(df_b)
    now_time = datetime.datetime.now().strftime('%H:%M:%S')

