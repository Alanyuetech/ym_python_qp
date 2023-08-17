# -*- coding: utf-8 -*-
"""
Created on Tue Oct 12 12:04:12 2021

@author: Administrator
"""

def wimi_bishu(wimi_bs,vena_bs,holo_bs):    
    "WIMI——笔数  生成"
    
    import akshare as ak
    import pandas as pd
    import time 
    from matplotlib import pyplot as plt 
    import numpy as np
    import datetime
    from dateutil.relativedelta import relativedelta
    from openpyxl import load_workbook
    import xlsxwriter
    import openpyxl
    from openpyxl.styles import PatternFill,Font,Alignment
    
    
    def getLastWeekDay(day=datetime.datetime.now()):
        now=day
        if now.isoweekday()==1:
          dayStep=3
        else:
          dayStep=1
        lastWorkDay = now - datetime.timedelta(days=dayStep)
        return lastWorkDay
    date_work=getLastWeekDay().strftime('%Y%m%d')
    
    
    df_info = pd.DataFrame(columns=('idd','ftnn_bs'))
    df_info.loc[0,'idd'] = 'WIMI'
    df_info.loc[1,'idd'] = 'MLGO'
    df_info.loc[2,'idd'] = 'HOLO'
    # df_info.loc[3,'idd'] = 'VENAR'
    df_info.loc[0,'ftnn_bs'] = wimi_bs
    df_info.loc[1,'ftnn_bs'] = vena_bs
    df_info.loc[2,'ftnn_bs'] = holo_bs
    # df_info.loc[3,'ftnn_bs'] = venar_bs
    
    
    
    def zongbishu(idd,ftnn_bs):
    
        "读取整理好的同花顺数据"
        ans_a = pd.read_excel('工作簿{}.xlsx'.format(idd),engine='openpyxl',dtype={'时间':'str'})
        
        ans_a['笔数'] = ans_a['笔数'].astype('str').str.replace('0','1').astype('int')
        ans_a['现手'] = ans_a['现手'].astype('str').str.replace(r'[^0-9]','') .astype('int')
        ans_a['笔数*系数'] = ans_a['笔数']*(ftnn_bs/ans_a['笔数'].sum())
        
        ans_a.to_excel("{}{}--笔数.xlsx".format(idd,date_work))
        
        
        workbook=openpyxl.load_workbook("{}{}--笔数.xlsx".format(idd,date_work))
        worksheet=workbook.worksheets[0]
        for i in range(2,len(ans_a)):
            worksheet["F{}".format(i)].number_format = '###'
        
        worksheet.delete_cols(1,1)
        workbook.save(filename="{}{}--笔数.xlsx".format(idd,date_work))
        print(idd)
        
        
    for index,row in df_info.iterrows():
        zongbishu(row['idd'],row['ftnn_bs'])
    
    ##############################################################################
    # "同花顺数据只有时间"
    # ans_a = pd.read_excel('工作簿.xlsx'.format(date_work),engine='openpyxl',dtype={'时间':'str'})
    # ans_a['笔数'] = 1
    # ans_a['笔数*系数'] = ans_a['笔数']*(ftnn_bs/ans_a['笔数'].sum())
    # ans_a.to_excel("WIMI{}--笔数.xlsx".format(date_work))
    
    # workbook=openpyxl.load_workbook("WIMI{}--笔数.xlsx".format(date_work))
    # worksheet=workbook.worksheets[0]
    # for i in range(2,len(ans_a)):
    #     worksheet["F{}".format(i)].number_format = '###'
    
    # worksheet.delete_cols(1,1)
    # workbook.save(filename="WIMI{}--笔数.xlsx".format(date_work))
    
if __name__ == '__main__':
    wimi_bs = 1322
    vena_bs = 8
    holo_bs = 4
    # venar_bs = 175
    wimi_bishu(wimi_bs,vena_bs,holo_bs)
    # wimi_bishu(wimi_bs,vena_bs,holo_bs,venar_bs)
    pass    
































